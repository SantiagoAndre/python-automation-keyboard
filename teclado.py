from pynput.keyboard import Key, Controller 
from time import sleep
import openpyxl
excel_document = openpyxl.load_workbook('notas.xlsx')
sheet = excel_document['aa']
keyboard  =Controller()
sleep(2)
print("start")
nota = sheet.cell(row = 1, column = 1).value	
posicion = 2
ant  = False
while True:
	if not nota:	
		if ant:
			break
		else:
			ant = True
	elif ant:
		ant = False
	for char in nota.__str__():
		keyboard.press(char)
		keyboard.release(char)
	keyboard.press(Key.enter)
	keyboard.release(Key.enter)	
	sleep(0.1)
	nota = sheet.cell(row = posicion, column = 1).value
	posicion+=1
